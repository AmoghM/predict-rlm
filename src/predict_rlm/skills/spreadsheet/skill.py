"""Spreadsheet skill — build, transform, and style .xlsx workbooks in the sandbox."""

from pathlib import Path

from predict_rlm import Skill

_MODULES_DIR = Path(__file__).parent / "modules"

spreadsheet_skill = Skill(
    name="spreadsheet",
    instructions="""# Spreadsheet Skill

## Workflow
1. Prefer `openpyxl` for `.xlsx` editing and formatting. Use `pandas` for analysis and CSV/TSV workflows.
2. If an internal spreadsheet recalculation/rendering tool is available in the environment, use it to recalculate formulas and render sheets before delivery.
3. Use formulas for derived values instead of hardcoding results.
4. If layout matters, render the output and inspect it visually with your sub-LM.
5. Save outputs, keep filenames stable, and clean up intermediate files.

## Primary tooling
- Use `openpyxl` for creating/editing `.xlsx` files and preserving formatting.
- Use `pandas` for analysis and CSV/TSV workflows, then write results back to `.xlsx` or `.csv`.
- Use `openpyxl.chart` for native Excel charts when needed.
- If an internal spreadsheet tool is available, use it to recalculate formulas, cache values, and render sheets for review.
- For sheets over ~1000 rows, read via `iter_rows(values_only=True)` with `read_only=True`; never iterate `ws.cell(r, c).value` in Python loops (prohibitively slow on WASM sandboxes, and still slow on native openpyxl for very large sheets).
- When writing Python-computed values to cells, match the expected data type: if the task calls for a numeric rank, count, or integer index derived from parsing a label string (e.g. `"item_3"` → `3`), extract and write `int()` — not the raw label string. Example: `cell.value = int(label.rsplit("_", 1)[-1])`.

## Reading efficiently (find a value without scanning everything)
- Read metadata before data. Use any workbook map provided in the instruction (sheet names, dimensions, header rows) — or `ws.max_row`/`ws.max_column` plus the header row — to locate a target's region first, then read only that region.
- Always BOUND your reads: pass `min_row`/`max_row` (and `min_col`/`max_col`) to `iter_rows`. Do not scan an entire sheet to find a value. Worksheets frequently report an inflated used-range (formatting applied to whole columns), so an unbounded `iter_rows(values_only=True)` can iterate ~1,000,000 phantom rows and stall the run.
- Do not use `ws.cell(row, col)` random access on a `read_only=True` worksheet — it re-reads from the top and is slow. Read the bounded range with `iter_rows`, or open just that sheet without `read_only` if you genuinely need random access.
- Locate each requested field once via the map/headers, read its bounded range, and move on. If a field is not present in the mapped structure, report it as not present — do not repeatedly re-scan or re-load the workbook hunting for it.

## Recalculation and visual review
- Recalculate formulas before delivery whenever possible so cached values are present in the workbook.
- `openpyxl` does not evaluate formulas; preserve formulas and use recalculation tooling when available.
- A `recalculate(path)` tool is available — call it on your output before submitting, then reload with `data_only=True` to confirm cells resolved.
- After reloading with `data_only=True`, inspect every target cell in the "Answer position" range: (a) if any cell `.value` is a formula-error string (`'#N/A'`, `'#REF!'`, `'#VALUE!'`, `'#DIV/0!'`, `'#NAME?'`, `'#NULL!'`), the formula failed — fix and re-save before submitting; (b) if any target cell `.value` is `None` when you intended to write a value, the write was silently dropped — re-open, re-write, and re-save.

## Rendering and visual checks
- Render the output sheet to a PNG when layout matters — call `render(path)` for the whole active sheet, or `render(path, cell_range="J3:N5")` / `render(path, cell_range="Sheet1!J3:N5")` to focus on a specific region. Use the cell_range form when your target area is on a wide/tall sheet and wouldn't fit on the first printed page.
- Inspect the rendered image with your sub-LM: `await predict("rendered: dspy.Image, task_instruction: str -> matches_instruction: bool, issues: list[str]", rendered=img_uri, task_instruction=task_instruction)` — pass `img_uri` as a `data:image/png;base64,...` URI.
- Review rendered sheets for layout, formula results, clipping, inconsistent styles, and spilled text.

## Formula requirements
- Use formulas for derived values rather than hardcoding results.
- Do not use dynamic array functions like `FILTER`, `XLOOKUP`, `SORT`, or `SEQUENCE`.
- Keep formulas simple and legible; use helper cells for complex logic.
- Avoid volatile functions like `INDIRECT` and `OFFSET` unless required.
- Prefer cell references over magic numbers (for example, `=H6*(1+$B$3)` instead of `=H6*1.04`).
- Use absolute (`$B$4`) or relative (`B4`) references carefully so copied formulas behave correctly.
- If you need literal text that starts with `=`, prefix it with a single quote.
- Guard against `#REF!`, `#DIV/0!`, `#VALUE!`, `#N/A`, and `#NAME?` errors.
- Check for off-by-one mistakes, circular references, and incorrect ranges.
- Prefix these functions with `_xlfn.` (e.g. `=_xlfn.IFNA(...)`) or they evaluate to `#NAME?`: `IFNA`, `IFS`, `SWITCH`, `TEXTJOIN`, `AGGREGATE`, `MAXIFS`, `MINIFS`, `CONCAT`.
- When a formula should produce an integer result (period counts, lookup indices, whole-unit financial figures), wrap it with `=ROUND(expr,0)` or `=INT(expr)` to avoid floating-point drift against integer reference values. Example: `=ROUND(DB($C$4,$C$5,$C$6,C8,1),0)` instead of a bare `DB()` call when reference values are whole numbers.

## Citation requirements
- When extracting or reporting values FROM a workbook in your answer, cite the source location of each value as `Sheet!Cell` (or `Sheet!Range`) so every extracted fact is traceable back to the cell it came from. Treat this as mandatory for data-extraction tasks: never state an extracted figure without its source sheet and cell/range.
- Cite sources inside the spreadsheet using plain-text URLs.
- For financial models, cite model inputs in cell comments.
- For tabular data sourced externally, add a source column when each row represents a separate item.

## Formatting requirements (existing formatted spreadsheets)
- Render and inspect a provided spreadsheet before modifying it when possible.
- Preserve existing formatting and style exactly.
- Match styles for any newly filled cells that were previously blank.
- Never overwrite established formatting unless the user explicitly asks for a redesign.

## Formatting requirements (new or unstyled spreadsheets)
- Use appropriate number and date formats.
- Dates should render as dates, not plain numbers.
- Percentages should usually default to one decimal place unless the data calls for something else.
- Currencies should use the appropriate currency format.
- Headers should be visually distinct from raw inputs and derived cells.
- Use fill colors, borders, spacing, and merged cells sparingly and intentionally.
- Set row heights and column widths so content is readable without excessive whitespace.
- Do not apply borders around every filled cell.
- Group related calculations and make totals simple sums of the cells above them.
- Add whitespace to separate sections.
- Ensure text does not spill into adjacent cells.
- Avoid unsupported spreadsheet data-table features such as `=TABLE`.

## Color conventions (if no style guidance)
- Blue: user input
- Black: formulas and derived values
- Green: linked or imported values
- Gray: static constants
- Orange: review or caution
- Light red: error or flag
- Purple: control or logic
- Teal: visualization anchors and KPI highlights

## Finance-specific requirements
- Format zeros as `-`.
- Negative numbers should be red and in parentheses.
- Format multiples as `5.2x`.
- Always specify units in headers (for example, `Revenue ($mm)`).
- Cite sources for all raw inputs in cell comments.
- For new financial models with no user-specified style, use blue text for hardcoded inputs, black for formulas, green for internal workbook links, red for external links, and yellow fill for key assumptions that need attention.

## Investment banking layouts
If the spreadsheet is an IB-style model (LBO, DCF, 3-statement, valuation):
- Totals should sum the range directly above.
- Hide gridlines and use horizontal borders above totals across relevant columns.
- Section headers should be merged cells with dark fill and white text.
- Column labels for numeric data should be right-aligned; row labels should be left-aligned.
- Indent submetrics under their parent line items.

## Additional Pitfalls

### VBA / Macro Task Requests
If the instruction asks for Visual Basic or macro code, **do not write VBA source text into cells**. openpyxl cannot embed executable VBA in an `.xlsx` file. Instead, identify the intended data transformation (sorting, filtering, row deletion, extraction, etc.) and perform it directly in Python. Write the resulting computed values or reorganized data into the target cell range.

### Date Values: Use datetime Objects, Not Formatted Strings
Assign `datetime.datetime` (or `datetime.date`) objects to `cell.value` when writing dates. openpyxl serializes Python date objects as Excel serial-date numbers; the evaluator reads them back as `datetime.datetime`. A string like `'01-01-2023'` is stored as text and will not match a datetime comparison.
Example: `cell.value = datetime(2023, 1, 1)` (correct) vs `cell.value = '01-01-2023'` (wrong — stored as text).

### Instruction-as-Question → Compute the Answer, Don't Explain
When an instruction is phrased as 'how can I…' or 'is there a method to…', the target cells contain **computed values** (scalars or formulas), not explanatory prose. Inspect the target range and the existing data; write results, not descriptions of how to achieve them.

### Formulas That Cache as None
If `cell.value` is `None` after loading with `data_only=True`, the formula result was never cached. Call `recalculate(output_path)` first, then re-open with `load_workbook(output_path, data_only=True)`. If the cell still reads as `None`, compute the value in Python and write it as a scalar.

### Missing Output Sheets
If the 'Answer position' names a sheet that does not exist in the workbook, create it before writing: `wb.create_sheet('SheetName')`. Never assume all referenced sheets are already present.
""",
    packages=["openpyxl", "pandas", "formulas"],
    modules={"formula_eval": str(_MODULES_DIR / "formula_eval.py")},
    # Transparently memoize read-only ``openpyxl.load_workbook`` so a model that
    # re-loads the same workbook many times across code blocks pays the parse
    # cost once. Pure runtime optimization — no instruction/behavior change —
    # installed at sandbox init so the model's existing
    # ``from openpyxl import load_workbook`` picks up the cached function.
    #
    # ONLY read-only loads are cached: read-only openpyxl workbooks are immutable,
    # so sharing one across calls is safe. Writable loads (the edit/save path) and
    # any error pass straight through, untouched. The cache key includes the file
    # mtime, so a workbook rewritten on disk (e.g. after ``recalculate``) is
    # re-parsed rather than served stale.
    setup="""
import os as _os
import openpyxl as _openpyxl

if not getattr(_openpyxl, "_rlm_cache_installed", False):
    _rlm_real_load_workbook = _openpyxl.load_workbook
    _rlm_wb_cache = {}

    def _rlm_cached_load_workbook(filename, **kwargs):
        if not kwargs.get("read_only", False):
            return _rlm_real_load_workbook(filename, **kwargs)
        try:
            key = (
                str(filename),
                _os.path.getmtime(filename),
                bool(kwargs.get("data_only", False)),
                bool(kwargs.get("keep_vba", False)),
                bool(kwargs.get("keep_links", True)),
                bool(kwargs.get("rich_text", False)),
            )
        except OSError:
            return _rlm_real_load_workbook(filename, **kwargs)
        wb = _rlm_wb_cache.get(key)
        if wb is None:
            wb = _rlm_real_load_workbook(filename, **kwargs)
            _rlm_wb_cache[key] = wb
        return wb

    _openpyxl.load_workbook = _rlm_cached_load_workbook
    _openpyxl._rlm_wb_cache = _rlm_wb_cache
    _openpyxl._rlm_cache_installed = True
""",
)
